"""
Phase 3: Reports Service
Handles report generation, scheduling, and metrics aggregation
"""

from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
import json
from io import BytesIO
import logging

from sqlalchemy import text, desc, func
from sqlalchemy.orm import Session
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

logger = logging.getLogger(__name__)


class ReportsService:
    """Service for report generation and scheduling"""

    def __init__(self, db: Session):
        self.db = db
        self.scheduler = BackgroundScheduler()
        if not self.scheduler.running:
            self.scheduler.start()

    def create_report_template(
        self,
        template_name: str,
        description: str,
        report_type: str,
        template_config: Dict
    ) -> Dict:
        """Create a new report template"""
        try:
            query = text("""
                INSERT INTO report_templates
                (template_name, description, report_type, template_config, is_active)
                VALUES (:name, :desc, :type, :config, true)
                RETURNING id, template_name, report_type
            """)
            result = self.db.execute(query, {
                "name": template_name,
                "desc": description,
                "type": report_type,
                "config": json.dumps(template_config)
            })
            row = result.fetchone()
            return {
                "template_id": row[0],
                "template_name": row[1],
                "report_type": row[2],
                "status": "created"
            }
        except Exception as e:
            logger.error(f"Error creating report template: {e}")
            raise

    def generate_pdf_report(
        self,
        template_id: int,
        findings: List[Dict]
    ) -> Tuple[BytesIO, int]:
        """Generate PDF report from findings"""
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )

            elements = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1f4788'),
                spaceAfter=30
            )
            elements.append(Paragraph("Database Analysis Report", title_style))
            elements.append(Spacer(1, 0.3 * inch))

            # Summary
            summary_data = [
                ["Metric", "Value"],
                ["Total Findings", str(len(findings))],
                ["Critical", str(sum(1 for f in findings if f.get("severity") == "Critical"))],
                ["High", str(sum(1 for f in findings if f.get("severity") == "High"))],
                ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            ]
            summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 0.3 * inch))

            # Findings table
            if findings:
                findings_data = [["ID", "Severity", "Issue", "Status"]]
                for finding in findings[:20]:  # Limit to 20 per page
                    findings_data.append([
                        str(finding.get("id", "")),
                        finding.get("severity", ""),
                        finding.get("issue", "")[:50],
                        finding.get("status", "")
                    ])

                findings_table = Table(findings_data, colWidths=[0.8*inch, 1*inch, 2.7*inch, 0.8*inch])
                findings_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
                ]))
                elements.append(findings_table)

            doc.build(elements)
            buffer.seek(0)
            return buffer, len(findings)

        except Exception as e:
            logger.error(f"Error generating PDF report: {e}")
            raise

    def generate_excel_report(
        self,
        template_id: int,
        findings: List[Dict]
    ) -> Tuple[BytesIO, int]:
        """Generate Excel report from findings"""
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Findings"

            # Headers
            headers = ["ID", "Severity", "Issue", "Status", "Created"]
            worksheet.append(headers)

            # Format header row
            header_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Add findings
            for finding in findings:
                worksheet.append([
                    finding.get("id", ""),
                    finding.get("severity", ""),
                    finding.get("issue", ""),
                    finding.get("status", ""),
                    finding.get("created", "")
                ])

            # Adjust column widths
            worksheet.column_dimensions['A'].width = 8
            worksheet.column_dimensions['B'].width = 12
            worksheet.column_dimensions['C'].width = 40
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 15

            # Summary sheet
            summary_sheet = workbook.create_sheet("Summary")
            summary_sheet.append(["Metric", "Value"])
            summary_sheet.append(["Total Findings", len(findings)])
            summary_sheet.append(["Critical", sum(1 for f in findings if f.get("severity") == "Critical")])
            summary_sheet.append(["High", sum(1 for f in findings if f.get("severity") == "High")])
            summary_sheet.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            return buffer, len(findings)

        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            raise

    def schedule_report(
        self,
        template_id: int,
        cron_expression: str,
        recipients: List[str],
        format_type: str
    ) -> Dict:
        """Schedule a report with cron expression"""
        try:
            query = text("""
                INSERT INTO scheduled_reports
                (template_id, cron_expression, recipients, format, is_active, next_run_at)
                VALUES (:tid, :cron, :recipients, :format, true, NOW() + INTERVAL '1 day')
                RETURNING id, schedule_name, next_run_at
            """)
            result = self.db.execute(query, {
                "tid": template_id,
                "cron": cron_expression,
                "recipients": ",".join(recipients),
                "format": format_type
            })
            row = result.fetchone()

            # Register scheduler job
            job_id = f"report_{row[0]}"
            trigger = CronTrigger.from_crontab(cron_expression)
            self.scheduler.add_job(
                self._execute_scheduled_report,
                trigger=trigger,
                id=job_id,
                args=[row[0]]
            )

            return {
                "scheduled_report_id": row[0],
                "schedule_name": row[1],
                "next_run_at": str(row[2]),
                "status": "scheduled"
            }
        except Exception as e:
            logger.error(f"Error scheduling report: {e}")
            raise

    def _execute_scheduled_report(self, scheduled_report_id: int):
        """Execute a scheduled report (internal use)"""
        try:
            query = text("""
                SELECT template_id, recipients, format
                FROM scheduled_reports
                WHERE id = :id
            """)
            result = self.db.execute(query, {"id": scheduled_report_id})
            row = result.fetchone()

            if row:
                template_id, recipients, format_type = row
                findings = self.get_findings_for_report()

                if format_type == "pdf":
                    buffer, count = self.generate_pdf_report(template_id, findings)
                else:
                    buffer, count = self.generate_excel_report(template_id, findings)

                # Log execution
                log_query = text("""
                    INSERT INTO report_executions
                    (scheduled_report_id, execution_date, row_count, execution_time_ms, status)
                    VALUES (:id, NOW(), :count, 0, 'success')
                """)
                self.db.execute(log_query, {"id": scheduled_report_id, "count": count})
                self.db.commit()

                logger.info(f"Report {scheduled_report_id} executed successfully")
        except Exception as e:
            logger.error(f"Error executing scheduled report: {e}")

    def get_findings_for_report(self) -> List[Dict]:
        """Retrieve findings for report generation"""
        try:
            query = text("""
                SELECT id, severity, issue, status, created_at
                FROM findings
                WHERE created_at >= NOW() - INTERVAL '30 days'
                ORDER BY severity DESC, created_at DESC
                LIMIT 500
            """)
            results = self.db.execute(query)
            findings = []
            for row in results:
                findings.append({
                    "id": row[0],
                    "severity": row[1],
                    "issue": row[2],
                    "status": row[3],
                    "created": str(row[4])
                })
            return findings
        except Exception as e:
            logger.error(f"Error retrieving findings: {e}")
            return []

    def get_report_execution_history(
        self,
        scheduled_report_id: int,
        limit: int = 20
    ) -> List[Dict]:
        """Get execution history for a scheduled report"""
        try:
            query = text("""
                SELECT id, execution_date, row_count, status
                FROM report_executions
                WHERE scheduled_report_id = :id
                ORDER BY execution_date DESC
                LIMIT :limit
            """)
            results = self.db.execute(query, {
                "id": scheduled_report_id,
                "limit": limit
            })
            history = []
            for row in results:
                history.append({
                    "execution_id": row[0],
                    "execution_date": str(row[1]),
                    "row_count": row[2],
                    "status": row[3]
                })
            return history
        except Exception as e:
            logger.error(f"Error retrieving execution history: {e}")
            return []

    def calculate_report_metrics(self, report_date: Optional[str] = None) -> Dict:
        """Calculate aggregated metrics for reports"""
        try:
            if report_date is None:
                report_date = datetime.now().strftime("%Y-%m-%d")

            query = text("""
                SELECT
                    COUNT(*) as total,
                    SUM(CASE WHEN severity = 'Critical' THEN 1 ELSE 0 END) as critical,
                    SUM(CASE WHEN severity = 'High' THEN 1 ELSE 0 END) as high,
                    SUM(CASE WHEN severity = 'Medium' THEN 1 ELSE 0 END) as medium,
                    SUM(CASE WHEN severity = 'Low' THEN 1 ELSE 0 END) as low,
                    SUM(CASE WHEN status = 'Resolved' THEN 1 ELSE 0 END) as resolved
                FROM findings
                WHERE DATE(created_at) = :date
            """)
            result = self.db.execute(query, {"date": report_date})
            row = result.fetchone()

            if row:
                total = row[0] or 0
                resolved = row[5] or 0
                resolution_rate = (resolved / total * 100) if total > 0 else 0

                metrics = {
                    "total_findings": total,
                    "critical_findings": row[1] or 0,
                    "high_findings": row[2] or 0,
                    "medium_findings": row[3] or 0,
                    "low_findings": row[4] or 0,
                    "findings_resolved": resolved,
                    "resolution_rate": round(resolution_rate, 2)
                }

                # Store metrics
                insert_query = text("""
                    INSERT INTO report_metrics
                    (report_date, total_findings, critical_findings, high_findings,
                     medium_findings, low_findings, findings_resolved)
                    VALUES (:date, :total, :critical, :high, :medium, :low, :resolved)
                    ON CONFLICT (report_date) DO UPDATE SET
                        total_findings = :total,
                        critical_findings = :critical,
                        high_findings = :high,
                        medium_findings = :medium,
                        low_findings = :low,
                        findings_resolved = :resolved
                """)
                self.db.execute(insert_query, {
                    "date": report_date,
                    "total": total,
                    "critical": row[1] or 0,
                    "high": row[2] or 0,
                    "medium": row[3] or 0,
                    "low": row[4] or 0,
                    "resolved": resolved
                })
                self.db.commit()

                return metrics
            return {}
        except Exception as e:
            logger.error(f"Error calculating metrics: {e}")
            return {}

    def get_finding_trends(self, days: int = 30) -> List[Dict]:
        """Get trend data for findings over time"""
        try:
            query = text("""
                SELECT date, severity, count
                FROM finding_trends
                WHERE date >= NOW()::date - :days::interval
                ORDER BY date DESC, severity
            """)
            results = self.db.execute(query, {"days": f"{days} days"})
            trends = []
            for row in results:
                trends.append({
                    "date": str(row[0]),
                    "severity": row[1],
                    "count": row[2]
                })
            return trends
        except Exception as e:
            logger.error(f"Error retrieving trends: {e}")
            return []

    def record_trend_data(self):
        """Record daily trend snapshot"""
        try:
            query = text("""
                INSERT INTO finding_trends (date, severity, count, cumulative_count)
                SELECT
                    NOW()::date,
                    severity,
                    COUNT(*),
                    (SELECT COUNT(*) FROM findings WHERE severity = f.severity)
                FROM findings f
                WHERE DATE(f.created_at) = NOW()::date
                GROUP BY f.severity
            """)
            self.db.execute(query)
            self.db.commit()
        except Exception as e:
            logger.error(f"Error recording trend data: {e}")

    def shutdown(self):
        """Shutdown the scheduler"""
        if self.scheduler.running:
            self.scheduler.shutdown()
