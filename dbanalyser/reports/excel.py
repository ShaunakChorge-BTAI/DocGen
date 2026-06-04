"""Excel report generator — writes a multi-sheet .xlsx workbook."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import pandas as pd

logger = logging.getLogger(__name__)

# Severity sort order
_SEV_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}


def _severity_sort(df: pd.DataFrame) -> pd.DataFrame:
    if "severity" in df.columns:
        df = df.copy()
        df["_sev_ord"] = df["severity"].map(_SEV_ORDER).fillna(9)
        df = df.sort_values("_sev_ord").drop(columns=["_sev_ord"])
    return df


def _auto_width(ws, df: pd.DataFrame) -> None:
    """Set approximate column widths from data."""
    try:
        from openpyxl.utils import get_column_letter
        for i, col in enumerate(df.columns, start=1):
            max_len = max(
                len(str(col)),
                df[col].astype(str).str.len().max() if not df.empty else 0,
            )
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 80)
    except Exception:
        pass  # openpyxl width setting is optional


def _header_style(ws, n_cols: int) -> None:
    """Bold + coloured header row."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.font       = font
            cell.fill       = fill
            cell.alignment  = Alignment(horizontal="center", vertical="center")
    except Exception:
        pass


def generate_excel(
        result,          # AnalysisResult
        output_path: str,
        dmv_results: Optional[dict] = None,
) -> str:
    """
    Write a multi-sheet Excel workbook.

    Returns the absolute path of the written file.
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    def _add_sheet(name: str, df: pd.DataFrame) -> None:
        if df is None or df.empty:
            ws = wb.create_sheet(title=name[:31])
            ws.append(["No data"])
            return
        ws = wb.create_sheet(title=name[:31])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        _header_style(ws, len(df.columns))
        _auto_width(ws, df)

    # ── Sheet 1: Executive Summary ──────────────────────────────────────────
    sev  = result.severity_counts
    summary_data = {
        "Metric": [
            "Run Label", "Source Mode", "Total Objects", "Total Findings",
            "Critical", "High", "Medium", "Low", "Overall Health Score", "Elapsed (s)",
        ],
        "Value": [
            result.run_label, result.source_mode,
            result.total_objects, result.total_findings,
            sev.get("Critical", 0), sev.get("High", 0),
            sev.get("Medium", 0),   sev.get("Low", 0),
            result.overall_health,  result.elapsed_sec,
        ],
    }
    _add_sheet("Summary", pd.DataFrame(summary_data))

    # ── Sheet 2: All Findings ───────────────────────────────────────────────
    all_df = result.all_findings_df()
    _add_sheet("All Findings", _severity_sort(all_df))

    # ── Sheet 3-N: Per-category sheets ─────────────────────────────────────
    if not all_df.empty:
        for cat in sorted(all_df["category"].unique()):
            cat_df = all_df[all_df["category"] == cat]
            _add_sheet(cat[:31], _severity_sort(cat_df))

    # ── Object health scores ────────────────────────────────────────────────
    health_rows = [
        {
            "object":        f"{r.obj.schema}.{r.obj.name}",
            "type":          r.obj.obj_type,
            "lines":         r.obj.lines,
            "health_score":  r.health_score,
            "critical":      r.severity_counts.get("Critical", 0),
            "high":          r.severity_counts.get("High", 0),
            "medium":        r.severity_counts.get("Medium", 0),
            "low":           r.severity_counts.get("Low", 0),
            "total_findings":len(r.findings),
        }
        for r in result.object_results
    ]
    _add_sheet("Object Health", pd.DataFrame(health_rows))

    # ── Extended checks ─────────────────────────────────────────────────────
    ext_name_map = {
        "tables_without_pk":      "Tables No PK",
        "duplicate_indexes":      "Duplicate Indexes",
        "column_type_mismatches": "Col Type Mismatches",
    }
    for key, sheet_name in ext_name_map.items():
        df = result.extended.get(key, pd.DataFrame())
        _add_sheet(sheet_name, df)

    # ── DMV sheets ──────────────────────────────────────────────────────────
    if dmv_results:
        dmv_name_map = {
            "dmv_index_usage":    "DMV - Index Usage",
            "dmv_missing_indexes":"DMV - Missing Indexes",
            "dmv_slow_queries":   "DMV - Slow Queries",
            "dmv_wait_stats":     "DMV - Wait Stats",
            "dmv_blocking_chains":"DMV - Blocking",
            "dmv_table_sizes":    "DMV - Table Sizes",
        }
        for key, sheet_name in dmv_name_map.items():
            df = dmv_results.get(key, pd.DataFrame())
            _add_sheet(sheet_name, df)

    # ── Save ─────────────────────────────────────────────────────────────────
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    logger.info("Excel report written to %s", path)
    return str(path)
