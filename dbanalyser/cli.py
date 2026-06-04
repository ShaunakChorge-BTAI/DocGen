"""
DBAnalyser CLI  (v2 — multi-DB)
================================
Entry point: `dbanalyser` (registered via pyproject.toml).

Commands
--------
  run        — Analyse SQL objects (single DB, named DB, or all registered DBs)
  report     — Generate reports from a stored run
  api        — Start the REST API server (FastAPI + uvicorn)
  validate   — Test connections (SQL Server + PostgreSQL)
  init-db    — Create / migrate the PostgreSQL schema
  history    — List previous analysis runs
  diff       — Compare two runs (new / resolved findings)
  db         — Manage the database registry
    db list      — Show all registered databases
    db add       — Add / update a database entry
    db remove    — Soft-delete (deactivate) a database entry
    db sync      — Sync analysis_config.yaml databases into PostgreSQL registry
    db show      — Show details + latest run stats for one database
"""

from __future__ import annotations

import logging
import sys
import time
from pathlib import Path
from typing import Optional

import click
from rich.console import Console
from rich.table import Table
from rich import box
from rich.progress import Progress, SpinnerColumn, TextColumn

console = Console()
logger  = logging.getLogger("dbanalyser")

_DEFAULT_CONFIG = Path(__file__).parent.parent / "analysis_config.yaml"
_SCHEMA_SQL     = Path(__file__).parent / "db" / "schema.sql"


# ─── Shared helpers ───────────────────────────────────────────────────────────

def _config_option(f):
    return click.option(
        "--config", "-c",
        default=str(_DEFAULT_CONFIG),
        show_default=True,
        help="Path to analysis_config.yaml",
    )(f)


def _setup_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        format="%(asctime)s  %(levelname)-8s  %(name)s — %(message)s",
        datefmt="%H:%M:%S",
        level=level,
    )


def _sev_colour(sev: int) -> str:
    if sev >= 5:   return "[red]"
    if sev >= 3:   return "[orange1]"
    if sev >= 1:   return "[yellow]"
    return "[green]"


def _health_style(score) -> str:
    if score is None: return "[dim]—[/dim]"
    s = float(score)
    if s < 50:  return f"[red]{s}[/red]"
    if s < 70:  return f"[orange1]{s}[/orange1]"
    if s < 85:  return f"[yellow]{s}[/yellow]"
    return f"[green]{s}[/green]"


# ─── Main group ───────────────────────────────────────────────────────────────

@click.group()
@click.version_option(version="2.0.0", prog_name="DBAnalyser")
@click.option("--verbose", "-v", is_flag=True, default=False)
@click.pass_context
def main(ctx: click.Context, verbose: bool) -> None:
    """DBAnalyser — enterprise SQL Server code quality & performance analyser."""
    ctx.ensure_object(dict)
    ctx.obj["verbose"] = verbose
    _setup_logging(verbose)


# ─── dbanalyser run ───────────────────────────────────────────────────────────

@main.command("run")
@_config_option
@click.option("--label",    "-l", default="",  help="Human-readable label for this run")
@click.option("--output-dir","-o", default="./output", show_default=True)
@click.option("--format",  "-f", "fmt",
              type=click.Choice(["all","excel","html","csv","json"], case_sensitive=False),
              default="all", show_default=True)
@click.option("--no-persist", is_flag=True, default=False,
              help="Skip writing to PostgreSQL")
@click.option("--dmv", is_flag=True, default=False,
              help="Run live-DB DMV analysis (requires live_db mode)")
@click.option("--db-name",  default=None,
              help="Run against a specific registered database (from databases: list)")
@click.option("--all-dbs",  is_flag=True, default=False,
              help="Run against every active database in the registry")
@click.pass_context
def cmd_run(ctx, config, label, output_dir, fmt,
            no_persist, dmv, db_name, all_dbs) -> None:
    """Run analysis and generate reports."""
    from .config  import load_config
    from .engine  import run_analysis
    from .reports import generate_excel, generate_html, generate_csv, generate_json

    cfg = load_config(config)

    # Build the list of database entries to iterate over
    if all_dbs:
        db_entries = cfg.get_active_databases()
        if not db_entries:
            console.print("[yellow]No active databases found in analysis_config.yaml.[/yellow]")
            console.print("Add entries under the 'databases:' key, then run [cyan]dbanalyser db sync[/cyan].")
            return
    elif db_name:
        entry = cfg.get_database(db_name)
        if not entry:
            console.print(f"[red]Database '{db_name}' not found in config. "
                          f"Available: {[d.name for d in cfg.databases]}[/red]")
            sys.exit(1)
        db_entries = [entry]
    else:
        db_entries = []  # single run against the top-level source config

    if db_entries:
        # ── Multi-DB run ────────────────────────────────────────────────────
        console.print(f"\n[bold]Running analysis across {len(db_entries)} database(s)[/bold]\n")
        summary_rows = []
        for db in db_entries:
            console.rule(f"[bold cyan]{db.name}[/bold cyan]  ({db.environment})")
            db_cfg    = cfg.settings_for_database(db)
            run_label = label or f"{db.name}_{time.strftime('%Y%m%d_%H%M%S')}"
            result    = _run_single(db_cfg, run_label, output_dir, fmt,
                                    no_persist, dmv, db_entry=db)
            summary_rows.append((db.name, db.environment,
                                  result.overall_health,
                                  result.severity_counts.get("Critical", 0),
                                  result.severity_counts.get("High", 0),
                                  result.total_findings))

        # Summary table
        t = Table(title="Multi-DB Run Summary", box=box.ROUNDED)
        t.add_column("Database",    style="bold")
        t.add_column("Env")
        t.add_column("Health",      justify="right")
        t.add_column("Critical",    justify="right", style="red")
        t.add_column("High",        justify="right", style="orange1")
        t.add_column("Total",       justify="right")
        for name, env, h, crit, high, total in summary_rows:
            t.add_row(name, env, str(h), str(crit), str(high), str(total))
        console.print(t)
    else:
        # ── Single run ──────────────────────────────────────────────────────
        run_label = label or time.strftime("%Y%m%d_%H%M%S")
        _run_single(cfg, run_label, output_dir, fmt, no_persist, dmv)


def _run_single(cfg, run_label: str, output_dir: str, fmt: str,
                no_persist: bool, dmv: bool, db_entry=None):
    """Run analysis for a single source config. Returns AnalysisResult."""
    from .engine  import run_analysis
    from .reports import generate_excel, generate_html, generate_csv, generate_json

    with Progress(SpinnerColumn(), TextColumn("{task.description}"),
                  transient=True) as prog:
        task = prog.add_task(f"Analysing {run_label} …", total=None)
        result = run_analysis(cfg, run_label=run_label)
        prog.update(task, description="Generating reports …")

    sev = result.severity_counts
    t   = Table(title=f"✓ {run_label}", box=box.SIMPLE)
    t.add_column("Metric",   style="bold")
    t.add_column("Value",    justify="right")
    t.add_row("Objects",     str(result.total_objects))
    t.add_row("Findings",    str(result.total_findings))
    t.add_row("[red]Critical",    str(sev.get("Critical", 0)))
    t.add_row("[orange1]High",    str(sev.get("High", 0)))
    t.add_row("[yellow]Medium",   str(sev.get("Medium", 0)))
    t.add_row("[green]Low",       str(sev.get("Low", 0)))
    t.add_row("Health",      _health_style(result.overall_health))
    t.add_row("Elapsed",     f"{result.elapsed_sec}s")
    console.print(t)

    # DMV
    dmv_results = None
    if dmv and cfg.source.mode == "live_db":
        from .engine.dmv import run_all_dmv_checks
        try:
            dmv_results = run_all_dmv_checks(cfg.source.connection_string)
            console.print(f"  [green]DMV checks complete[/green]")
        except Exception as exc:
            console.print(f"  [yellow]DMV skipped — {exc}[/yellow]")

    # Reports
    out  = Path(output_dir)
    ts   = run_label.replace(":", "-")
    written = []
    if fmt in ("all", "excel"):
        written.append(generate_excel(result, str(out / f"dbanalyser_{ts}.xlsx"), dmv_results))
    if fmt in ("all", "html"):
        written.append(generate_html(result, str(out / f"dbanalyser_{ts}.html"), dmv_results))
    if fmt in ("all", "csv"):
        written.extend(generate_csv(result, str(out / ts), dmv_results))
    if fmt in ("all", "json"):
        written.append(generate_json(result, str(out / f"dbanalyser_{ts}.json"), dmv_results))
    for p in written[:6]:
        console.print(f"  [cyan]→[/cyan] {p}")
    if len(written) > 6:
        console.print(f"  [dim]… and {len(written)-6} more files[/dim]")

    # Persist
    if not no_persist:
        try:
            _persist(cfg, result, run_label, dmv_results, db_entry)
        except Exception as exc:
            console.print(f"[yellow]Postgres persist skipped — {exc}[/yellow]")

    return result


def _persist(cfg, result, run_label: str, dmv_results, db_entry=None) -> None:
    from .db.connection  import init_pool, close_pool
    from .db.repository  import (insert_run, bulk_insert_findings,
                                   upsert_db_registry, update_db_registry_last_run,
                                   upsert_health_trend, get_new_vs_resolved,
                                   detect_and_mark_content_drift,
                                   enrich_findings_with_history)
    from .db.models      import DbRegistry, Finding, HealthTrend, Run
    import uuid

    init_pool(cfg.postgres)
    try:
        # Resolve db_registry_id
        db_registry_id = None
        if db_entry:
            reg = DbRegistry(
                name          = db_entry.name,
                environment   = db_entry.environment,
                host          = db_entry.host,
                port          = db_entry.port,
                database_name = db_entry.database_name,
                connection_string = db_entry.connection_string or None,
                use_windows_auth  = db_entry.use_windows_auth,
                username          = db_entry.username or None,
                password          = db_entry.password or None,
                description       = db_entry.description or None,
                owner_label       = db_entry.owner_label or None,
                tags              = list(db_entry.tags),
                is_active         = db_entry.is_active,
            )
            db_registry_id = upsert_db_registry(reg)

        sev = result.severity_counts
        run = Run(
            run_id         = str(uuid.uuid4()),
            label          = run_label,
            db_registry_id = db_registry_id,
            source_mode    = result.source_mode,
            total_objects  = result.total_objects,
            total_issues   = result.total_findings,
            critical_count = sev.get("Critical", 0),
            high_count     = sev.get("High",     0),
            medium_count   = sev.get("Medium",   0),
            low_count      = sev.get("Low",      0),
            health_score   = result.overall_health,
            status         = "success",
        )
        run_int_id = insert_run(run)

        # Findings
        findings = [
            Finding(
                run_id         = run_int_id,
                schema_name    = or_.obj.schema,
                object_name    = or_.obj.name,
                object_type    = or_.obj.obj_type,
                rule_id        = f.rule_id or "",
                category       = f.category,
                severity       = f.severity,
                issue          = f.issue,
                recommendation = f.recommendation,
                line_number    = f.line_number,
                snippet        = f.snippet or "",
            )
            for or_ in result.object_results
            for f in or_.findings
        ]
        bulk_insert_findings(run_int_id, findings)

        # Drift detection + findings deduplication (post-insert)
        try:
            drifted = detect_and_mark_content_drift(run_int_id, db_registry_id)
            deduped = enrich_findings_with_history(run_int_id, db_registry_id)
            if drifted or deduped:
                console.print(
                    f"  [dim]Drift: {drifted} changed objects | "
                    f"Dedup: {deduped} repeat findings suppressed[/dim]"
                )
        except Exception as exc:
            console.print(f"  [yellow]Warning: drift/dedup step failed: {exc}[/yellow]")

        # Trend row
        trend = HealthTrend(
            run_id         = run_int_id,
            db_registry_id = db_registry_id,
            db_name        = db_entry.name if db_entry else cfg.source.file_path,
            environment    = cfg.run.environment,
            health_score   = result.overall_health,
            total_objects  = result.total_objects,
            total_issues   = result.total_findings,
            critical_count = sev.get("Critical", 0),
            high_count     = sev.get("High",     0),
            medium_count   = sev.get("Medium",   0),
            low_count      = sev.get("Low",      0),
        )
        upsert_health_trend(trend)

        # Update registry last_run
        if db_registry_id:
            update_db_registry_last_run(db_registry_id, result.overall_health)

        console.print(
            f"  [green]Persisted {len(findings)} findings "
            f"(run_id={run_int_id})[/green]"
        )
    finally:
        close_pool()


# ─── dbanalyser db (sub-group) ────────────────────────────────────────────────

@main.group("db")
def db_group() -> None:
    """Manage the database registry (multi-DB Phase 1)."""


@db_group.command("list")
@_config_option
@click.option("--all", "show_all", is_flag=True, default=False,
              help="Include inactive databases")
@click.pass_context
def db_list(ctx, config, show_all) -> None:
    """List all registered databases and their latest health scores."""
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .db.repository import get_db_summary

    cfg = load_config(config)
    try:
        init_pool(cfg.postgres)
        rows = get_db_summary()
        close_pool()
        if not rows:
            console.print("[yellow]No databases registered yet.[/yellow]")
            console.print("Run [cyan]dbanalyser db sync[/cyan] to import from config.")
            return
        t = Table(title="Registered Databases", box=box.ROUNDED)
        t.add_column("Name",        style="bold cyan")
        t.add_column("Env")
        t.add_column("Owner")
        t.add_column("Health",      justify="right")
        t.add_column("Critical",    justify="right", style="red")
        t.add_column("High",        justify="right", style="orange1")
        t.add_column("Findings",    justify="right")
        t.add_column("Last Run")
        t.add_column("Active")
        for r in rows:
            if not show_all and not r.get("is_active"):
                continue
            t.add_row(
                r["name"],
                r.get("environment", ""),
                r.get("owner_label") or "—",
                _health_style(r.get("health_score")),
                str(r.get("critical_count") or 0),
                str(r.get("high_count") or 0),
                str(r.get("total_issues") or 0),
                str(r.get("last_run_ts") or "Never")[:19],
                "[green]✓[/green]" if r.get("is_active") else "[dim]✗[/dim]",
            )
        console.print(t)
    except Exception as exc:
        console.print(f"[yellow]Could not load from PostgreSQL ({exc}). "
                      f"Showing config file databases instead.[/yellow]")
        cfg = load_config(config)
        for db in cfg.databases:
            console.print(f"  {'[green]✓[/green]' if db.is_active else '[dim]✗[/dim]'}  "
                          f"[cyan]{db.name}[/cyan]  {db.environment}  {db.owner_label or '—'}")


@db_group.command("sync")
@_config_option
@click.pass_context
def db_sync(ctx, config) -> None:
    """Sync the databases: list from config into the PostgreSQL registry."""
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .db.repository import upsert_db_registry
    from .db.models     import DbRegistry

    cfg = load_config(config)
    if not cfg.databases:
        console.print("[yellow]No databases defined in analysis_config.yaml.[/yellow]")
        return

    init_pool(cfg.postgres)
    try:
        for db in cfg.databases:
            reg = DbRegistry(
                name          = db.name,
                environment   = db.environment,
                host          = db.host,
                port          = db.port,
                database_name = db.database_name,
                connection_string = db.connection_string or None,
                use_windows_auth  = db.use_windows_auth,
                username          = db.username or None,
                password          = db.password or None,
                description       = db.description or None,
                owner_label       = db.owner_label or None,
                tags              = list(db.tags),
                is_active         = db.is_active,
            )
            db_id = upsert_db_registry(reg)
            status = "[green]✓[/green]" if db.is_active else "[dim]inactive[/dim]"
            console.print(f"  {status}  {db.name}  (id={db_id})")
        console.print(f"\n[green]Synced {len(cfg.databases)} database(s).[/green]")
    finally:
        close_pool()


@db_group.command("add")
@_config_option
@click.argument("name")
@click.option("--env",    default="development", show_default=True)
@click.option("--host",   default="localhost",   show_default=True)
@click.option("--port",   default=1433,          show_default=True, type=int)
@click.option("--db",     "database_name", default="",
              help="SQL Server database name")
@click.option("--dsn",    "connection_string", default="",
              help="Full pyodbc connection string (overrides host/port/db)")
@click.option("--owner",  default="", help="Owner / team label")
@click.option("--desc",   default="", help="Description")
@click.pass_context
def db_add(ctx, config, name, env, host, port, database_name,
           connection_string, owner, desc) -> None:
    """Add or update a database in the registry."""
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .db.repository import upsert_db_registry
    from .db.models     import DbRegistry

    cfg = load_config(config)
    init_pool(cfg.postgres)
    try:
        reg = DbRegistry(
            name=name, environment=env, host=host, port=port,
            database_name=database_name,
            connection_string=connection_string or None,
            use_windows_auth=False,
            username="",
            password=None,
            description=desc or None, owner_label=owner or None,
        )
        db_id = upsert_db_registry(reg)
        console.print(f"[green]✓ '{name}' registered (id={db_id})[/green]")
    finally:
        close_pool()


@db_group.command("remove")
@_config_option
@click.argument("db_id", type=int)
@click.pass_context
def db_remove(ctx, config, db_id) -> None:
    """Deactivate (soft-delete) a database from the registry by ID."""
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .db.repository import delete_db_registry

    cfg = load_config(config)
    init_pool(cfg.postgres)
    try:
        found = delete_db_registry(db_id)
        if found:
            console.print(f"[green]Database id={db_id} deactivated.[/green]")
        else:
            console.print(f"[yellow]Database id={db_id} not found.[/yellow]")
    finally:
        close_pool()


@db_group.command("show")
@_config_option
@click.argument("name")
@click.pass_context
def db_show(ctx, config, name) -> None:
    """Show details and last-run stats for a specific database."""
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .db.repository import get_db_registry, list_runs

    cfg = load_config(config)
    init_pool(cfg.postgres)
    try:
        row = get_db_registry(name)
        if not row:
            console.print(f"[red]'{name}' not found in registry.[/red]")
            return
        console.print(f"\n[bold cyan]{row['name']}[/bold cyan]")
        for k in ("environment","host","port","database_name",
                  "owner_label","description","is_active","tags",
                  "last_run_at","last_health"):
            console.print(f"  {k:<20}: {row.get(k, '—')}")

        runs = list_runs(limit=5, db_registry_id=row["id"])
        if runs:
            console.print("\n  [bold]Last 5 runs:[/bold]")
            t = Table(box=box.SIMPLE)
            t.add_column("Label")
            t.add_column("Health", justify="right")
            t.add_column("Findings", justify="right")
            t.add_column("Timestamp")
            for r in runs:
                t.add_row(r.get("label",""),
                          _health_style(r.get("health_score")),
                          str(r.get("total_issues",0)),
                          str(r.get("timestamp",""))[:19])
            console.print(t)
    finally:
        close_pool()


# ─── dbanalyser report ────────────────────────────────────────────────────────

@main.command("report")
@_config_option
@click.option("--run-id",    default=None, type=int,
              help="Integer run ID (default: latest)")
@click.option("--db-name",   default=None,
              help="Scope to latest run for this database")
@click.option("--output-dir","-o", default="./output", show_default=True)
@click.option("--format",    "-f", "fmt",
              type=click.Choice(["excel","html","csv","json"], case_sensitive=False),
              default="excel", show_default=True)
@click.pass_context
def cmd_report(ctx, config, run_id, db_name, output_dir, fmt) -> None:
    """Generate a report from a stored run (reads from PostgreSQL)."""
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .db.repository import get_run, get_findings, get_db_registry

    cfg = load_config(config)
    init_pool(cfg.postgres)
    try:
        if db_name:
            reg = get_db_registry(db_name)
            if not reg:
                console.print(f"[red]Database '{db_name}' not in registry.[/red]")
                return
            run_row = get_run()  # latest
        else:
            run_row = get_run(run_id=run_id)

        if not run_row:
            console.print("[red]No run found.[/red]")
            return

        console.print(f"Generating {fmt} report for run: {run_row.get('label','')}")
        findings_df = get_findings(run_row["id"])
        console.print(f"  Loaded {len(findings_df)} findings from PostgreSQL")

        out = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)
        ts  = str(run_row.get("label","run")).replace(":", "-")
        if fmt == "excel":
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            wb = Workbook(); ws = wb.active; ws.title = "Findings"
            if not findings_df.empty:
                for r in dataframe_to_rows(findings_df, index=False, header=True):
                    ws.append(r)
            path = out / f"dbanalyser_{ts}.xlsx"
            wb.save(str(path))
        elif fmt == "json":
            import json
            path = out / f"dbanalyser_{ts}.json"
            path.write_text(
                json.dumps(findings_df.fillna("").to_dict(orient="records"), indent=2),
                encoding="utf-8")
        elif fmt == "csv":
            path = out / f"dbanalyser_{ts}_findings.csv"
            findings_df.to_csv(path, index=False)
        elif fmt == "html":
            path = out / f"dbanalyser_{ts}.html"
            path.write_text(findings_df.to_html(index=False), encoding="utf-8")
        console.print(f"  [cyan]→[/cyan] {path}")
    finally:
        close_pool()


# ─── dbanalyser validate ──────────────────────────────────────────────────────

@main.command("validate")
@_config_option
@click.option("--db-name", default=None,
              help="Validate a specific registered database connection")
@click.pass_context
def cmd_validate(ctx, config, db_name) -> None:
    """Test connections (SQL Server + PostgreSQL)."""
    from .config import load_config

    cfg = load_config(config)
    console.print("[bold]Validating connections …[/bold]\n")

    # Decide which SQL Server entries to test
    if db_name:
        entry = cfg.get_database(db_name)
        sql_entries = [entry] if entry else []
    else:
        sql_entries = cfg.get_active_databases()

    if not sql_entries and cfg.source.connection_string:
        # Fall back to the top-level source.connection_string
        class _FakeEntry:
            name = "source"
            effective_connection_string = cfg.source.connection_string
        sql_entries = [_FakeEntry()]

    # Test SQL Server connections
    for entry in sql_entries:
        try:
            import pyodbc  # type: ignore
            conn = pyodbc.connect(entry.effective_connection_string, timeout=10)
            ver  = conn.execute("SELECT @@VERSION").fetchone()[0].split("\n")[0]
            conn.close()
            console.print(f"  [green]✓[/green] SQL Server [{entry.name}]: {ver}")
        except Exception as exc:
            console.print(f"  [red]✗[/red] SQL Server [{entry.name}]: {exc}")

    # Test PostgreSQL
    try:
        from .db.connection import test_connection
        ok, msg = test_connection(cfg.postgres)
        sym = "[green]✓[/green]" if ok else "[red]✗[/red]"
        console.print(f"  {sym} PostgreSQL: {msg}")
    except Exception as exc:
        console.print(f"  [red]✗[/red] PostgreSQL: {exc}")


# ─── dbanalyser init-db ───────────────────────────────────────────────────────

@main.command("init-db")
@_config_option
@click.pass_context
def cmd_init_db(ctx, config) -> None:
    """Create / migrate the PostgreSQL schema for DBAnalyser."""
    from .config        import load_config
    from .db.connection import create_schema

    cfg = load_config(config)
    console.print("Creating PostgreSQL schema …")
    try:
        create_schema(cfg.postgres, str(_SCHEMA_SQL))
        console.print("  [green]✓ Schema created / verified.[/green]")
        console.print("\n  Next step: [cyan]dbanalyser db sync[/cyan]  "
                      "(imports databases from config)")
    except Exception as exc:
        console.print(f"  [red]✗ Failed: {exc}[/red]")
        sys.exit(1)


# ─── dbanalyser history ───────────────────────────────────────────────────────

@main.command("history")
@_config_option
@click.option("--limit",   default=20,  show_default=True)
@click.option("--db-name", default=None,
              help="Filter to a specific registered database")
@click.pass_context
def cmd_history(ctx, config, limit, db_name) -> None:
    """List previous analysis runs."""
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .db.repository import list_runs, get_db_registry

    cfg = load_config(config)
    init_pool(cfg.postgres)
    try:
        db_id = None
        if db_name:
            reg = get_db_registry(db_name)
            if not reg:
                console.print(f"[red]'{db_name}' not in registry.[/red]")
                return
            db_id = reg["id"]

        runs = list_runs(limit=limit, db_registry_id=db_id)
        if not runs:
            console.print("[yellow]No runs found.[/yellow]")
            return

        t = Table(title="Analysis Run History", box=box.SIMPLE)
        t.add_column("ID",       justify="right")
        t.add_column("Label")
        t.add_column("Database")
        t.add_column("Mode")
        t.add_column("Objects",  justify="right")
        t.add_column("Findings", justify="right")
        t.add_column("Health",   justify="right")
        t.add_column("Timestamp")
        for r in runs:
            t.add_row(
                str(r["id"]),
                r.get("label",""),
                r.get("db_name") or "—",
                r.get("source_mode",""),
                str(r.get("total_objects",0)),
                str(r.get("total_issues",0)),
                _health_style(r.get("health_score")),
                str(r.get("timestamp",""))[:19],
            )
        console.print(t)
    finally:
        close_pool()


# ─── dbanalyser diff ──────────────────────────────────────────────────────────

@main.command("diff")
@_config_option
@click.argument("run_id_a", type=int)
@click.argument("run_id_b", type=int)
@click.pass_context
def cmd_diff(ctx, config, run_id_a: int, run_id_b: int) -> None:
    """Compare two runs — show new / resolved findings. RUN_ID_A is the baseline."""
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .db.repository import get_findings
    import pandas as pd

    cfg = load_config(config)
    init_pool(cfg.postgres)
    try:
        df_a = get_findings(run_id_a)
        df_b = get_findings(run_id_b)
        if isinstance(df_a, list): df_a = pd.DataFrame(df_a)
        if isinstance(df_b, list): df_b = pd.DataFrame(df_b)

        key = ["object_name", "rule_id"]
        set_a = set(df_a[key].apply(tuple, axis=1)) if not df_a.empty else set()
        set_b = set(df_b[key].apply(tuple, axis=1)) if not df_b.empty else set()
        new_keys      = set_b - set_a
        resolved_keys = set_a - set_b

        console.print(f"\n[bold]Diff: Run {run_id_a} → Run {run_id_b}[/bold]")
        console.print(f"  [green]+{len(new_keys)} new[/green]   "
                      f"[red]-{len(resolved_keys)} resolved[/red]   "
                      f"{len(set_a & set_b)} unchanged\n")

        for title, keys, df in [
            ("NEW Findings",      new_keys,      df_b),
            ("RESOLVED Findings", resolved_keys, df_a),
        ]:
            if not keys: continue
            mask = df.apply(lambda r: (r["object_name"], r["rule_id"]) in keys, axis=1)
            sub  = df[mask].head(30)
            if sub.empty: continue
            t = Table(title=title, box=box.SIMPLE_HEAD)
            for col in ("object_name","rule_id","severity","issue"):
                if col in sub.columns:
                    t.add_column(col.replace("_"," ").title())
            for _, row in sub.iterrows():
                t.add_row(*[str(row[c]) for c in ("object_name","rule_id","severity","issue")
                            if c in sub.columns])
            console.print(t)
    finally:
        close_pool()


# ─── dbanalyser api ───────────────────────────────────────────────────────────

@main.command("api")
@_config_option
@click.option("--host",    default="0.0.0.0", show_default=True,
              help="Bind host for the API server")
@click.option("--port",    default=8000, show_default=True, type=int,
              help="TCP port for the API server")
@click.option("--reload",  is_flag=True, default=False,
              help="Enable auto-reload (development mode)")
@click.option("--api-key", default=None, envvar="DBANALYSER_API_KEY",
              help="Require X-API-Key header (or ?api_key=). "
                   "If omitted the API is open (local use only).")
@click.pass_context
def cmd_api(ctx, config, host, port, reload, api_key) -> None:
    """Start the DBAnalyser REST API server (FastAPI + uvicorn).

    \b
    Swagger UI:   http://<host>:<port>/docs
    ReDoc:        http://<host>:<port>/redoc
    Health check: http://<host>:<port>/health
    """
    try:
        from dbanalyser.api.main import start_api
    except ImportError as exc:
        console.print(f"[red]FastAPI / uvicorn not installed: {exc}[/red]")
        console.print("Run: [cyan]pip install 'dbanalyser[api]'[/cyan]  "
                      "or  [cyan]pip install fastapi uvicorn[standard][/cyan]")
        sys.exit(1)

    console.print(
        f"[bold cyan]DBAnalyser REST API[/bold cyan]  "
        f"-> http://{host}:{port}/docs"
    )
    if api_key:
        console.print(f"  [green]Auth enabled[/green] — X-API-Key required")
    else:
        console.print("  [yellow]Auth disabled[/yellow] — open access (local use only)")

    start_api(
        config_path=config,
        host=host,
        port=port,
        reload=reload,
        api_key=api_key,
    )


# ─── dbanalyser compliance-report ────────────────────────────────────────────

@main.command("compliance-report")
@_config_option
@click.option("--run-id",    default=None, type=int,
              help="Integer run ID (default: latest)")
@click.option("--output-dir","-o", default="./output", show_default=True)
@click.option("--format",    "-f", "fmt",
              type=click.Choice(["excel", "json", "csv"], case_sensitive=False),
              default="excel", show_default=True)
@click.option("--packs",    default="sox,gdpr,rbi,dng",
              help="Comma-separated compliance packs to include (sox,gdpr,rbi,dng)")
@click.pass_context
def cmd_compliance_report(ctx, config, run_id, output_dir, fmt, packs) -> None:
    """Generate a compliance-only report (SOX / GDPR / RBI / Dangerous DML).

    \b
    Examples:
      dbanalyser compliance-report
      dbanalyser compliance-report --run-id 5 --packs sox,gdpr --format excel
    """
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .db.repository import get_run, get_findings, list_runs
    import pandas as pd

    cfg = load_config(config)
    init_pool(cfg.postgres)
    try:
        run_row = get_run(run_id=run_id) if run_id else None
        if run_row is None:
            rows = list_runs(limit=1)
            run_row = rows[0] if rows else None
        if run_row is None:
            console.print("[red]No runs found in PostgreSQL. Run an analysis first.[/red]")
            return

        findings_df = get_findings(run_row["id"])
        if isinstance(findings_df, list):
            findings_df = pd.DataFrame(findings_df)

        console.print(
            f"Generating compliance report for run: "
            f"[cyan]{run_row.get('label', run_row['id'])}[/cyan]  "
            f"({len(findings_df)} total findings)"
        )

        # ── Filter to requested packs ─────────────────────────────────────
        PACK_PREFIXES = {
            "sox":  "Compliance-SOX",
            "gdpr": "Compliance-GDPR",
            "rbi":  "Compliance-RBI",
            "dng":  "Dangerous SQL",
        }
        requested = [p.strip().lower() for p in packs.split(",")]
        prefixes  = [PACK_PREFIXES[p] for p in requested if p in PACK_PREFIXES]

        if not prefixes:
            console.print("[red]No valid pack names specified.[/red]"); return

        comp_df = findings_df[
            findings_df["category"].isin(prefixes)
        ].copy() if not findings_df.empty and "category" in findings_df.columns else pd.DataFrame()

        console.print(f"  Compliance findings: [bold]{len(comp_df)}[/bold]")
        for prefix in prefixes:
            n = int((comp_df["category"] == prefix).sum()) if not comp_df.empty else 0
            console.print(f"    {prefix}: {n}")

        # ── Generate output ───────────────────────────────────────────────
        from pathlib import Path
        out = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)
        ts  = str(run_row.get("label", "run")).replace(":", "-")

        if fmt == "excel":
            try:
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                console.print("[red]openpyxl required: pip install openpyxl[/red]"); return

            wb = Workbook()
            # Summary sheet
            ws_sum = wb.active
            ws_sum.title = "Summary"
            ws_sum["A1"] = "DBAnalyser — Compliance Report"
            ws_sum["A1"].font = Font(bold=True, size=14)
            ws_sum["A2"] = f"Run: {run_row.get('label', run_row['id'])}"
            ws_sum["A3"] = f"Health Score: {run_row.get('health_score', '—')}"
            ws_sum["A5"] = "Pack"
            ws_sum["B5"] = "Findings"
            ws_sum["A5"].font = ws_sum["B5"].font = Font(bold=True)
            for i, prefix in enumerate(prefixes, 6):
                n = int((comp_df["category"] == prefix).sum()) if not comp_df.empty else 0
                ws_sum.cell(row=i, column=1, value=prefix)
                ws_sum.cell(row=i, column=2, value=n)

            # One sheet per pack
            for prefix in prefixes:
                sheet_name = prefix.replace("Compliance-", "").replace(" ", "_")[:30]
                ws = wb.create_sheet(title=sheet_name)
                sub = comp_df[comp_df["category"] == prefix] if not comp_df.empty else pd.DataFrame()
                if sub.empty:
                    ws["A1"] = "No findings for this pack."
                else:
                    for row in dataframe_to_rows(sub, index=False, header=True):
                        ws.append(row)
                    # Bold header
                    for cell in ws[1]:
                        cell.font = Font(bold=True)

            path = out / f"compliance_{ts}.xlsx"
            wb.save(str(path))

        elif fmt == "json":
            import json
            path = out / f"compliance_{ts}.json"
            payload = {
                "run": {k: str(v) for k, v in run_row.items()},
                "packs": {
                    prefix: comp_df[comp_df["category"] == prefix].fillna("").to_dict("records")
                    for prefix in prefixes
                } if not comp_df.empty else {},
            }
            path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")

        elif fmt == "csv":
            path = out / f"compliance_{ts}.csv"
            if comp_df.empty:
                pd.DataFrame().to_csv(path, index=False)
            else:
                comp_df.to_csv(path, index=False)

        console.print(f"  [cyan]→[/cyan] {path}")

    finally:
        close_pool()


# ─── dbanalyser schedule ──────────────────────────────────────────────────────

@main.group("schedule")
def schedule_group() -> None:
    """Manage scheduled analysis tasks."""


@schedule_group.command("list")
@_config_option
@click.pass_context
def schedule_list(ctx, config) -> None:
    """List all scheduled analysis tasks."""
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .scheduler     import list_tasks

    cfg = load_config(config)
    init_pool(cfg.postgres)
    try:
        tasks = list_tasks()
        if not tasks:
            console.print("[yellow]No scheduled tasks found.[/yellow]")
            console.print("Add one with: [cyan]dbanalyser schedule add <DB_NAME> --cron daily@02:00[/cyan]")
            return
        t = Table(title="Scheduled Tasks", box=box.ROUNDED)
        t.add_column("ID",       justify="right")
        t.add_column("Database", style="bold cyan")
        t.add_column("Schedule")
        t.add_column("Label")
        t.add_column("Enabled")
        t.add_column("Last Run")
        t.add_column("Next Run")
        for row in tasks:
            t.add_row(
                str(row.get("id", "")),
                row.get("db_name", ""),
                row.get("schedule", ""),
                row.get("label", "") or "—",
                "[green]✓[/green]" if row.get("enabled") else "[dim]✗[/dim]",
                str(row.get("last_run") or "Never")[:19],
                str(row.get("next_run") or "—")[:19],
            )
        console.print(t)
    finally:
        close_pool()


@schedule_group.command("add")
@_config_option
@click.argument("db_name")
@click.option("--cron",   "schedule", default="manual", show_default=True,
              help="Schedule: hourly | daily@HH:MM | weekly@DAY@HH:MM | manual")
@click.option("--label",  default="",  help="Run label (default: auto-generated)")
@click.option("--dmv",    "run_dmv", is_flag=True, default=False,
              help="Also run DMV checks")
@click.option("--formats", default="excel,json", show_default=True,
              help="Comma-separated report formats")
@click.pass_context
def schedule_add(ctx, config, db_name, schedule, label, run_dmv, formats) -> None:
    """Add or update a scheduled task for a registered database."""
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .scheduler     import ScheduledTask, add_task

    cfg = load_config(config)
    if not cfg.get_database(db_name):
        console.print(
            f"[red]Database '{db_name}' not found in config. "
            f"Available: {[d.name for d in cfg.databases]}[/red]"
        )
        return

    init_pool(cfg.postgres)
    try:
        task = ScheduledTask(
            db_name=db_name, schedule=schedule, label=label,
            run_dmv=run_dmv,
            formats=[f.strip() for f in formats.split(",")],
        )
        task_id = add_task(task)
        if task_id > 0:
            console.print(
                f"[green]✓ Scheduled task for '{db_name}' saved "
                f"(id={task_id}, schedule={schedule})[/green]"
            )
        else:
            console.print("[red]Failed to save scheduled task.[/red]")
    finally:
        close_pool()


@schedule_group.command("remove")
@_config_option
@click.argument("db_name")
@click.pass_context
def schedule_remove(ctx, config, db_name) -> None:
    """Remove a scheduled task for a database."""
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .scheduler     import remove_task

    cfg = load_config(config)
    init_pool(cfg.postgres)
    try:
        found = remove_task(db_name)
        if found:
            console.print(f"[green]Scheduled task for '{db_name}' removed.[/green]")
        else:
            console.print(f"[yellow]No scheduled task found for '{db_name}'.[/yellow]")
    finally:
        close_pool()


@schedule_group.command("run-due")
@_config_option
@click.pass_context
def schedule_run_due(ctx, config) -> None:
    """Execute all tasks that are past their scheduled time.

    \b
    Intended to be called frequently (e.g. every minute) from Windows Task
    Scheduler:  dbanalyser schedule run-due --config analysis_config.yaml
    """
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .scheduler     import run_due_tasks

    cfg = load_config(config)
    init_pool(cfg.postgres)
    try:
        n = run_due_tasks(config)
        if n:
            console.print(f"[green]Executed {n} scheduled task(s).[/green]")
        else:
            console.print("[dim]No tasks due.[/dim]")
    finally:
        close_pool()


# ─── dbanalyser auth ──────────────────────────────────────────────────────────

@main.group("auth")
def auth_group() -> None:
    """JWT auth utilities (hash passwords, generate tokens)."""


@auth_group.command("hash-password")
@click.argument("password", default="", required=False)
def auth_hash_password(password: str) -> None:
    """Generate a bcrypt hash for a password (for use in auth.users config).

    \b
    Examples:
      dbanalyser auth hash-password myS3cr3t
      dbanalyser auth hash-password          ← prompts for password
    """
    if not password:
        import getpass
        password = getpass.getpass("Password: ")
    try:
        from .api.auth_rbac import hash_password
        hashed = hash_password(password)
        console.print(f"[green]Password hash (add to auth.users.password_hash):[/green]")
        console.print(hashed)
    except ImportError:
        console.print("[red]bcrypt not installed: pip install bcrypt[/red]")


# ─── dbanalyser ingest ────────────────────────────────────────────────────────

@main.command("ingest")
@_config_option
@click.option("--db", default="", help="Database name from analysis_config.yaml databases list")
@click.option("--files", default="", help="Path to SQL files (overrides config source)")
@click.option("--db-registry-id", default=None, type=int,
              help="db_registry_id to tag ingested objects (auto-resolved from --db)")
@click.option("--use-transformers", is_flag=True, default=False,
              help="Use sentence-transformers for higher-quality embeddings (requires install)")
def cmd_ingest(config, db, files, db_registry_id, use_transformers) -> None:
    """Ingest schema into the vector knowledge base for AI optimization.

    \b
    Extracts schema objects (tables, columns, procedures, views, indexes)
    from a live SQL Server database or local .sql files, then computes
    embeddings and stores them in the schema_objects table.

    \b
    Examples:
      dbanalyser ingest --db LTFS_PROD
      dbanalyser ingest --files ./sql_scripts --db-registry-id 1
    """
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .schema_intel.extractor   import extract_schema_from_live_db, extract_schema_from_objects
    from .schema_intel.repository  import upsert_schema_object, get_schema_summary

    cfg = load_config(config)
    init_pool(cfg.postgres)

    try:
        db_entry = cfg.get_database(db) if db else None

        # Resolve db_registry_id
        reg_id = db_registry_id
        if reg_id is None and db_entry:
            try:
                from .db.connection import get_cursor
                with get_cursor() as cur:
                    cur.execute("SELECT id FROM db_registry WHERE name = %s", (db,))
                    row = cur.fetchone()
                    if row:
                        reg_id = row["id"]
            except Exception:
                pass

        # Extract schema objects
        if db_entry:
            console.print(f"[cyan]Extracting schema from live DB: {db}…[/cyan]")
            conn_str = db_entry.effective_connection_string
            objects  = extract_schema_from_live_db(conn_str)
        elif files:
            console.print(f"[cyan]Scanning SQL files in: {files}…[/cyan]")
            from .engine.scanner import scan_files
            sql_objects = list(scan_files(
                files,
                include_schemas=cfg.scope.schemas or None,
                include_types=cfg.scope.object_types or None,
            ))
            objects = extract_schema_from_objects(sql_objects)
        else:
            # Use config source
            src_path = cfg.source.file_path
            console.print(f"[cyan]Scanning config source: {src_path}…[/cyan]")
            from .engine.scanner import scan_files
            sql_objects = list(scan_files(
                src_path,
                include_schemas=cfg.scope.schemas or None,
                include_types=cfg.scope.object_types or None,
            ))
            objects = extract_schema_from_objects(sql_objects)

        if not objects:
            console.print("[yellow]No schema objects found — nothing to ingest.[/yellow]")
            return

        console.print(f"[green]Extracted {len(objects)} schema objects. Embedding…[/green]")

        ok = 0
        errors = 0
        with console.status("Upserting embeddings…"):
            for obj in objects:
                row_id = upsert_schema_object(reg_id, obj, use_transformers=use_transformers)
                if row_id >= 0:
                    ok += 1
                else:
                    errors += 1

        summary = get_schema_summary(reg_id)
        console.print(f"[green]Ingested {ok} objects ({errors} errors).[/green]")
        if summary:
            from rich.table import Table
            tbl = Table(title="Schema Knowledge Base Summary")
            tbl.add_column("Type"); tbl.add_column("Count", justify="right")
            for obj_type, count in sorted(summary.items()):
                tbl.add_row(obj_type, str(count))
            console.print(tbl)

    finally:
        close_pool()


# ─── dbanalyser optimize ──────────────────────────────────────────────────────

@main.command("optimize")
@_config_option
@click.argument("object_name")
@click.option("--sql-file", default="", help="Path to .sql file containing the object source")
@click.option("--execution-plan", default="", help="Path to XML execution plan file")
@click.option("--db-registry-id", default=None, type=int)
@click.option("--model", default="", help="Claude model ID (overrides config)")
@click.option("--no-persist", is_flag=True, default=False, help="Do not save result to DB")
def cmd_optimize(config, object_name, sql_file, execution_plan, db_registry_id,
                 model, no_persist) -> None:
    """AI-optimize a SQL object using Anthropic Claude.

    \b
    Fetches schema context from the knowledge base, then sends the SQL
    to Claude with the schema context, rule findings, and execution plan
    (if provided) for optimization suggestions.

    \b
    Examples:
      dbanalyser optimize dbo.usp_ProcessPayment --sql-file ./usp_ProcessPayment.sql
      dbanalyser optimize dbo.usp_X --sql-file x.sql --execution-plan plan.xml
    """
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .ai_optimizer  import build_optimization_context, optimize_sql_object

    cfg = load_config(config)
    init_pool(cfg.postgres)

    try:
        # Load SQL source
        if sql_file:
            with open(sql_file, encoding="utf-8", errors="replace") as f:
                source_sql = f.read()
        else:
            source_sql = click.edit(
                "-- Paste SQL source here\n",
                extension=".sql",
            ) or ""
            if not source_sql.strip():
                console.print("[red]No SQL source provided.[/red]")
                return

        # Load execution plan
        plan_text = ""
        if execution_plan:
            with open(execution_plan, encoding="utf-8", errors="replace") as f:
                plan_text = f.read()

        api_key = cfg.ai_optimizer.api_key or ""
        chosen_model = model or cfg.ai_optimizer.model

        console.print(f"[cyan]Building schema context for {object_name}…[/cyan]")
        ctx = build_optimization_context(
            object_name=object_name,
            source_sql=source_sql,
            db_registry_id=db_registry_id,
            execution_plan=plan_text,
        )

        for w in ctx["warnings"]:
            console.print(f"[yellow]Warning: {w}[/yellow]")

        console.print(
            f"[cyan]Calling Claude ({chosen_model}) — context quality: "
            f"{ctx['context_quality']}…[/cyan]"
        )

        result = optimize_sql_object(
            object_name=object_name,
            source_sql=source_sql,
            schema_context=ctx["schema_context"],
            findings=ctx["findings"],
            execution_plan=plan_text,
            api_key=api_key,
            model=chosen_model,
            db_registry_id=db_registry_id,
            persist=not no_persist,
        )

        if result.error:
            console.print(f"[red]Optimization failed: {result.error}[/red]")
            return

        console.print(
            f"\n[green]Optimization complete[/green]  "
            f"confidence={result.confidence_score:.0%}  "
            f"tokens={result.tokens_used:,}  "
            f"elapsed={result.elapsed_sec:.1f}s"
        )
        console.print("\n[bold]=== Reasoning ===[/bold]")
        console.print(result.reasoning)
        console.print("\n[bold]=== Optimized SQL ===[/bold]")
        console.print(result.optimized_sql)

    finally:
        close_pool()


# ─── dbanalyser audit ─────────────────────────────────────────────────────────

@main.command("audit")
@_config_option
@click.option("--username", default="", help="Filter by username")
@click.option("--action",   default="", help="Filter by action (e.g. optimize, ingest)")
@click.option("--limit",    default=50, type=int, help="Max rows to display")
def cmd_audit(config, username, action, limit) -> None:
    """View the audit log — history of user actions.

    \b
    Examples:
      dbanalyser audit
      dbanalyser audit --action optimize --limit 20
      dbanalyser audit --username alice
    """
    from .config        import load_config
    from .db.connection import init_pool, close_pool
    from .audit         import get_audit_logs

    cfg = load_config(config)
    init_pool(cfg.postgres)

    try:
        entries = get_audit_logs(
            username=username or None,
            action=action or None,
            limit=limit,
        )
        if not entries:
            console.print("[dim]No audit log entries found.[/dim]")
            return

        from rich.table import Table
        tbl = Table(title=f"Audit Log (last {limit})")
        tbl.add_column("Timestamp", no_wrap=True)
        tbl.add_column("User")
        tbl.add_column("Action")
        tbl.add_column("Resource Type")
        tbl.add_column("Resource")
        tbl.add_column("IP")
        for e in entries:
            tbl.add_row(
                str(e.created_at)[:19],
                e.username,
                e.action,
                e.resource_type,
                e.resource_id[:40],
                e.ip_address,
            )
        console.print(tbl)
    finally:
        close_pool()


# ─── Entry point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    main()
